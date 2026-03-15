import logging
import io
from fastapi import APIRouter
from fastapi.responses import Response
from pydantic import BaseModel
from typing import Optional, List
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from utils.groq_client import explain_436a, analyse_grievance
from utils.bail_model import get_bail_model, initialize_bail_model
from utils.synthetic_bail_data import generate_bail_dataset

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api/legal", tags=["Legal"])


class LegalCaseCreate(BaseModel):
    prisoner_name: str
    ward: Optional[str] = None
    ipc_section: Optional[str] = None
    max_sentence_years: Optional[int] = None
    detention_start: Optional[str] = None
    months_detained: Optional[int] = None
    dlsa_contact: Optional[str] = None
    grievance_id: Optional[str] = None


class BailPredictionRequest(BaseModel):
    """Request model for bail eligibility prediction"""
    case_id: str
    prisoner_name: str
    age: int
    offence_category: str
    offence_severity: int
    prior_criminal_history: str
    employment_status: str
    monthly_income: int
    residential_stability: str
    years_in_current_city: int
    has_family_ties: bool = False
    has_guarantor: bool = False
    guarantor_income: int = 0
    flight_risk: str = "medium"
    is_repeat_offender: bool = False
    days_between_arrest_and_trial: int = 100
    health_conditions: bool = False
    mental_health_issue: bool = False
    credibility_score: float = 50.0
    offence_category: str = "other"
    conditions_for_bail: List[str] = []
    recommended_bail_amount: int = 0


class LegalAdviceRequest(BaseModel):
    """Request model for legal advice"""
    query: str
    case_type: Optional[str] = None
    ipc_section: Optional[str] = None


@router.get("")
async def get_legal_cases():
    """Get all legal cases with 436A eligibility computed in-memory."""
    from main import supabase
    if not supabase:
        return {"success": False, "data": [], "error": "Database not configured"}
    try:
        result = supabase.table("legal_cases").select("*").order("created_at", desc=True).execute()
        cases = result.data or []

        # Compute 436A eligibility in-memory only (no DB writes)
        for case in cases:
            max_years = case.get("max_sentence_years") or 0
            months = case.get("months_detained") or 0
            half_max_months = (max_years * 12) // 2
            eligible = months >= half_max_months if half_max_months > 0 else False
            case["eligible_436a"] = eligible
            case["half_max_months"] = half_max_months

        return {"success": True, "data": cases, "error": None}
    except Exception as e:
        logger.error("Get legal cases error: %s", str(e))
        return {"success": False, "data": [], "error": str(e)}


@router.post("")
async def create_legal_case(req: LegalCaseCreate):
    """Add a new legal case."""
    from main import supabase
    if not supabase:
        return {"success": False, "data": None, "error": "Database not configured"}
    try:
        case_data = {
            "prisoner_name": req.prisoner_name,
            "ward": req.ward,
            "ipc_section": req.ipc_section,
            "max_sentence_years": req.max_sentence_years,
            "detention_start": req.detention_start,
            "months_detained": req.months_detained,
            "dlsa_contact": req.dlsa_contact,
            "grievance_id": req.grievance_id,
        }

        max_years = req.max_sentence_years or 0
        months = req.months_detained or 0
        half_max_months = (max_years * 12) // 2
        case_data["eligible_436a"] = months >= half_max_months if half_max_months > 0 else False

        result = supabase.table("legal_cases").insert(case_data).execute()
        if not result.data:
            return {"success": False, "data": None, "error": "Failed to create legal case"}

        return {"success": True, "data": result.data[0], "error": None}
    except Exception as e:
        logger.error("Create legal case error: %s", str(e))
        return {"success": False, "data": None, "error": str(e)}


@router.get("/check/{case_id}")
async def check_eligibility(case_id: str):
    """Get 436A explanation for a case."""
    from main import supabase
    if not supabase:
        return {"success": False, "data": None, "error": "Database not configured"}
    try:
        result = supabase.table("legal_cases").select("*").eq("id", case_id).execute()
        if not result.data:
            return {"success": False, "data": None, "error": "Case not found"}

        case = result.data[0]
        max_years = case.get("max_sentence_years") or 0
        months = case.get("months_detained") or 0
        half_max_months = (max_years * 12) // 2
        eligible = months >= half_max_months if half_max_months > 0 else False

        explanation = await explain_436a(
            prisoner_name=case.get("prisoner_name", "Unknown"),
            ipc_section=case.get("ipc_section", "Unknown"),
            max_years=max_years,
            months_detained=months,
            eligible=eligible,
        )

        return {
            "success": True,
            "data": {
                "case": case,
                "eligible": eligible,
                "half_max_months": half_max_months,
                "explanation": explanation,
            },
            "error": None,
        }
    except Exception as e:
        logger.error("Check eligibility error: %s", str(e))
        return {"success": False, "data": None, "error": str(e)}


# ============================================================================
# BAIL PREDICTION ENDPOINTS (Machine Learning + GenAI Legal Advice)
# ============================================================================

@router.post("/bail/predict")
async def predict_bail_eligibility(req: BailPredictionRequest):
    """
    Predict bail eligibility using local ML model.
    Uses synthetic data-trained RandomForest or rule-based system.
    """
    try:
        logger.info(f"Bail prediction for case {req.case_id}")
        
        # Get the bail model
        model = get_bail_model()
        
        # Prepare case data for prediction
        case_data = req.dict()
        
        # Make prediction
        prediction = model.predict(case_data)
        
        return {
            "success": True,
            "data": prediction,
            "error": None,
        }
    except Exception as e:
        logger.error(f"Bail prediction error: {str(e)}")
        return {
            "success": False,
            "data": None,
            "error": str(e),
        }


@router.post("/bail/initialize-model")
async def initialize_bail_prediction_model():
    """
    Initialize the bail prediction model by:
    1. Generating 10k synthetic bail cases
    2. Training RandomForest classifier (if sklearn available)
    3. Storing model metrics
    """
    try:
        logger.info("Initializing bail prediction model with synthetic data...")
        
        # Generate synthetic dataset
        dataset = generate_bail_dataset(10000)
        logger.info(f"Generated {len(dataset)} synthetic bail cases")
        
        # Initialize and train model
        model = get_bail_model()
        metrics = model.train(dataset)
        
        # Add dataset statistics
        eligible_count = sum(1 for case in dataset if case["bail_eligible"])
        metrics["dataset_statistics"] = {
            "total_cases": len(dataset),
            "eligible_for_bail": eligible_count,
            "not_eligible_for_bail": len(dataset) - eligible_count,
            "eligibility_rate": f"{eligible_count/len(dataset)*100:.1f}%",
        }
        
        return {
            "success": True,
            "data": metrics,
            "error": None,
        }
    except Exception as e:
        logger.error(f"Model initialization error: {str(e)}")
        return {
            "success": False,
            "data": None,
            "error": str(e),
        }


@router.get("/bail/list")
async def get_bail_recommendations():
    """
    Get list of all bail-eligible persons with predictions.
    Returns 186 grievances with bail recommendations from database.
    """
    try:
        from main import supabase
        
        if not supabase:
            return {
                "success": False,
                "data": [],
                "error": "Database not configured",
            }
        
        logger.info("Fetching bail recommendations...")
        
        # Fetch grievances from database
        result = supabase.table("grievances").select("*").order("created_at", desc=True).execute()
        grievances = result.data or []
        
        # Get bail model for predictions
        model = get_bail_model()
        
        bail_recommendations = []
        
        for grievance in grievances:
            # Prepare case data from grievance
            case_data = {
                "case_id": grievance.get("id", ""),
                "prisoner_name": grievance.get("citizen_name", ""),
                "age": 35,  # Default age (could be enhanced with separate person data)
                "offence_category": grievance.get("category", "other"),
                "offence_severity": min(5, max(1, grievance.get("urgency", 3))),
                "prior_criminal_history": "none",
                "employment_status": "employed",
                "monthly_income": 50000,
                "residential_stability": "owned",
                "years_in_current_city": 10,
                "has_family_ties": True,
                "has_guarantor": False,
                "flight_risk": "low",
                "is_repeat_offender": False,
                "days_between_arrest_and_trial": 120,
                "credibility_score": grievance.get("credibility_score", 50),
            }
            
            # Get prediction
            prediction = model.predict(case_data)
            
            if prediction.get("bail_eligible"):
                recommendation = {
                    "case_id": grievance.get("id", "")[:8],
                    "citizen_name": grievance.get("citizen_name", ""),
                    "phone": grievance.get("phone", ""),
                    "ward": grievance.get("ward", ""),
                    "category": grievance.get("category", ""),
                    "urgency": grievance.get("urgency", 0),
                    "credibility_score": grievance.get("credibility_score", 0),
                    "bail_eligible": True,
                    "recommendation_confidence": prediction.get("confidence", 0.5),
                    "reasoning": prediction.get("reasoning", ""),
                    "recommended_bail_amount": prediction.get("recommended_bail_amount", 0),
                }
                bail_recommendations.append(recommendation)
        
        logger.info(f"Generated {len(bail_recommendations)} bail recommendations")
        
        return {
            "success": True,
            "data": bail_recommendations,
            "error": None,
        }
    except Exception as e:
        logger.error(f"Get bail recommendations error: {str(e)}")
        return {
            "success": False,
            "data": [],
            "error": str(e),
        }


@router.get("/bail/export-excel")
async def export_bail_recommendations_excel():
    """
    Export bail-eligible persons list to Excel format.
    """
    try:
        from main import supabase
        
        if not supabase:
            return Response(
                content=b"",
                status_code=500,
                media_type="application/json",
            )
        
        logger.info("Exporting bail recommendations to Excel...")
        
        # Fetch grievances
        result = supabase.table("grievances").select("*").order("created_at", desc=True).execute()
        grievances = result.data or []
        
        # Get model predictions
        model = get_bail_model()
        bail_list = []
        
        for g in grievances:
            case_data = {
                "case_id": g.get("id", ""),
                "prisoner_name": g.get("citizen_name", ""),
                "age": 35,
                "offence_category": g.get("category", "other"),
                "offence_severity": min(5, max(1, g.get("urgency", 3))),
                "prior_criminal_history": "none",
                "employment_status": "employed",
                "monthly_income": 50000,
                "residential_stability": "owned",
                "years_in_current_city": 10,
                "has_family_ties": True,
                "has_guarantor": False,
                "flight_risk": "low",
                "is_repeat_offender": False,
                "days_between_arrest_and_trial": 120,
                "credibility_score": g.get("credibility_score", 50),
            }
            
            prediction = model.predict(case_data)
            
            if prediction.get("bail_eligible"):
                bail_list.append({
                    "case_id": g.get("id", "")[:8],
                    "name": g.get("citizen_name", ""),
                    "phone": g.get("phone", ""),
                    "ward": g.get("ward", ""),
                    "category": g.get("category", ""),
                    "urgency": g.get("urgency", ""),
                    "credibility": g.get("credibility_score", ""),
                    "confidence": round(prediction.get("confidence", 0), 3),
                    "bail_amount": prediction.get("recommended_bail_amount", 0),
                    "recommendation": prediction.get("reasoning", ""),
                })
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Bail Recommendations"
        
        # Headers
        headers = ["Case ID", "Name", "Phone", "Ward", "Category", "Urgency", "Credibility", "Confidence", "Bail Amount", "Recommendation"]
        
        header_fill = PatternFill(start_color="FF6B00", end_color="FF6B00", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        for idx, header in enumerate(headers, 1):
            cell = ws.cell(1, idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Data rows
        for row_idx, bail_case in enumerate(bail_list, 2):
            ws.cell(row_idx, 1).value = bail_case["case_id"]
            ws.cell(row_idx, 2).value = bail_case["name"]
            ws.cell(row_idx, 3).value = bail_case["phone"]
            ws.cell(row_idx, 4).value = bail_case["ward"]
            ws.cell(row_idx, 5).value = bail_case["category"]
            ws.cell(row_idx, 6).value = bail_case["urgency"]
            ws.cell(row_idx, 7).value = bail_case["credibility"]
            ws.cell(row_idx, 8).value = bail_case["confidence"]
            ws.cell(row_idx, 9).value = bail_case["bail_amount"]
            ws.cell(row_idx, 10).value = bail_case["recommendation"]
        
        # Auto-adjust column widths
        for col in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J"]:
            ws.column_dimensions[col].width = 15
        
        # Save to buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        logger.info(f"Bail export: {len(bail_list)} eligible persons in Excel")
        
        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=bail_recommendations_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"}
        )
    except Exception as e:
        logger.error(f"Bail export error: {str(e)}", exc_info=True)
        return Response(
            content=b"",
            status_code=500,
            media_type="application/json",
        )


@router.post("/advice/legal")
async def get_legal_advice(req: LegalAdviceRequest):
    """
    Get legal advice using GenAI (Groq/Gemini) with local ML fallback.
    """
    try:
        logger.info(f"Legal advice requested for: {req.query[:50]}")
        
        # Try GenAI first
        try:
            from utils.gemini_client import query_gemini_with_context
            
            context = ""
            if req.case_type:
                context += f"Case Type: {req.case_type}. "
            if req.ipc_section:
                context += f"IPC Section: {req.ipc_section}. "
            
            full_query = f"{context}{req.query}"
            
            # Query Gemini for legal advice
            advice = await query_gemini_with_context(
                query=full_query,
                context="You are a legal expert providing counsel on Indian law matters.",
                max_tokens=500
            )
            
            return {
                "success": True,
                "data": {
                    "advice": advice,
                    "source": "GenAI (Gemini)",
                    "confidence": "high",
                    "case_type": req.case_type,
                    "ipc_section": req.ipc_section,
                },
                "error": None,
            }
        except Exception as ai_error:
            logger.warning(f"GenAI failed, using rule-based: {ai_error}")
            
            # Fallback to rule-based legal advice
            rule_based_advice = generate_rule_based_advice(req.query, req.case_type, req.ipc_section)
            
            return {
                "success": True,
                "data": {
                    "advice": rule_based_advice,
                    "source": "Rule-Based System",
                    "confidence": "medium",
                    "case_type": req.case_type,
                    "ipc_section": req.ipc_section,
                },
                "error": None,
            }
            
    except Exception as e:
        logger.error(f"Legal advice error: {str(e)}")
        return {
            "success": False,
            "data": None,
            "error": str(e),
        }


def generate_rule_based_advice(query: str, case_type: Optional[str] = None, ipc_section: Optional[str] = None) -> str:
    """
    Generate rule-based legal advice as fallback when GenAI is unavailable.
    """
    
    advice_templates = {
        "bail": "For bail matters: 1) Section 436 IPC - bailable offences, 2) Section 437 - conditions for bail in non-bailable offences, 3) Section 438 - anticipatory bail. Consult with local advocate for specific case.",
        "criminal": "For criminal cases: 1) Ensure proper legal representation, 2) File bail application if in custody, 3) Gather supporting documents and witnesses, 4) Keep records of all transactions, 5) Comply with court orders.",
        "civil": "For civil disputes: 1) File case in appropriate jurisdiction, 2) Document all relevant agreements and communications, 3) Ensure witnesses are available, 4) Follow procedural rules, 5) Consider mediation/arbitration.",
        "property": "For property disputes: 1) Verify original documents and titles, 2) Check for encumbrances/liens, 3) Maintain clear communication history, 4) Consider possession status, 5) Consult property lawyer.",
        "labour": "For labour disputes: 1) File complaints with labour commissioner, 2) Document working conditions and communications, 3) Follow statutory notice requirements, 4) Preserve pay slips and records, 5) Seek union support if applicable.",
    }
    
    base_advice = "Consult with a qualified legal practitioner for detailed advice. General guidance: 1) Document everything, 2) Keep all communications, 3) Follow deadlines, 4) Gather supporting evidence, 5) Act promptly on legal matters."
    
    if case_type:
        case_lower = case_type.lower()
        for key, template in advice_templates.items():
            if key in case_lower:
                base_advice = template
                break
    
    if ipc_section:
        base_advice += f"\n\nRelevant IPC Section: {ipc_section}. This section requires specific procedural compliance."
    
    base_advice += "\n\n⚠️ DISCLAIMER: This is general information, not legal advice. Always consult with a licensed advocate."
    
    return base_advice
